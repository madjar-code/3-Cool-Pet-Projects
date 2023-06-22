from typing import List, Dict
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from django.conf import settings


def generate_excel_table(contacts: List[Dict[str, str]]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    sheet: Worksheet = wb.active

    headers = ['Name', 'Email', 'Phone', 'Priority Group']

    # Setting excel table headers
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell: Cell = sheet[f'{col_letter}1']
        cell.value = header
        cell.alignment = Alignment(horizontal='center',
                                   vertical='center',
                                   wrap_text=True)
        cell.font = cell.font.copy(bold=True)

    # Populating with data
    for row_num, contact in enumerate(contacts, 2):
        sheet.cell(row=row_num, column=1).value = contact['name']
        sheet.cell(row=row_num, column=2).value = contact['email']
        sheet.cell(row=row_num, column=3).value = contact['phone']
        sheet.cell(row=row_num, column=4).value = contact['priority_group']

    # Auto-fit table sizes
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].auto_size = True

    return wb


contacts_data: List[Dict[str, str]] = [
    {
        'name': 'John Doe',
        'email': 'john@example.com',
        'phone': '+88005553535',
        'priority_group': 'Low'
    },
    {
        'name': 'Jane Smith',
        'email': 'jane@example.com',
        'phone': '+88005553535',
        'priority_group': 'High'
    }
]

wb: openpyxl.Workbook = generate_excel_table(contacts_data)

wb.save('ENS/server/contacts.xlsx')