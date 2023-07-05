import os
from enum import Enum
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import (
    List,
    Dict,
)
from django.core.files import File
from rest_framework import status
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.parsers import (
    MultiPartParser,
)
from rest_framework.views import APIView
from rest_framework.generics import (
    GenericAPIView,
    ListAPIView,
    RetrieveAPIView,
    CreateAPIView,
    DestroyAPIView,
)
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from reports.models import (
    NotificationState,
    StateStatusChoices,
)
from contacts.models import Contact
from .serializers import (
    SimpleContactSerializer,
    ContactSerializer,
    UpdateContactSerializer,
    CreateContactSerializer,
)


class ErrorMessages(str, Enum):
    NO_USER = 'User with `username` not found'
    NO_CONTACT = 'Contact with `id` not found'
    INVALID_TYPE = 'Invalid file type. Only Excel files (.xlsx) are supported.'


class ContactsListView(ListAPIView):
    serializer_class = SimpleContactSerializer
    # permission_classes = (IsAdminUser,)

    queryset = Contact.active_objects.all()

    @swagger_auto_schema(operation_id='all_contacts')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


class ContactDetailsView(RetrieveAPIView):
    serializer_class = ContactSerializer
    # permission_classes = (IsAdminUser,)
    queryset = Contact.active_objects.all()
    lookup_field = 'id'

    @swagger_auto_schema(operation_id='contact_details')
    def get(self, request: Request, id: str) -> Response:
        contact: Contact = self.queryset.filter(id=id).first()
        if not contact:
            return Response({'error': ErrorMessages.NO_CONTACT.value},
                            status=status.HTTP_404_NOT_FOUND,)
        contact_notif_states = NotificationState.objects.filter(contact=contact)
        success_counter = contact_notif_states.filter(
            status=StateStatusChoices.STATUS_READY).count()
        failed_counter = contact_notif_states.filter(
            status=StateStatusChoices.STATUS_FAILED).count()

        serializer = self.serializer_class(instance=contact)
        return Response({'contact': serializer.data,
                         'notifications': {'success': success_counter,
                                           'failed': failed_counter}},
                        status.HTTP_200_OK)


class CreateContactView(CreateAPIView):
    serializer_class = CreateContactSerializer
    # permission_classes = (IsAdminUser,)

    @swagger_auto_schema(operation_id='create_contact')
    def post(self, request: Request) -> Response:
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class CreateMultiContactsView(CreateAPIView):
    serializer_class = CreateContactSerializer
    # permission_classes = (IsAdminUser,)

    @swagger_auto_schema(
        operation_id='create_multiply_contacts',
        request_body=openapi.Schema(
            type=openapi.TYPE_ARRAY,
            items=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    field: openapi.Schema(type=openapi.TYPE_STRING)
                    for field in serializer_class().fields.keys()
                }
            ),
            example=[
                {
                    'name': 'John',
                    'email': 'john@example.com',
                    'phone': '+12125552368',
                    'priority_group': 'Low',
                },
            ]
        )
    )
    def post(self, request: Request) -> Response:
        contacts = []
        errors = []

        for id, contact_data in enumerate(request.data):
            serializer = CreateContactSerializer(data=contact_data)
            if serializer.is_valid():
                if not self.check_email_in_request(id, contact_data, contacts):
                    errors.append({
                        'id': id,
                        'errors': [
                            {'email': ['This email is already in the request']}
                            ]
                        })
                contacts.append(serializer.data)
            else:
                errors.append({'id': id, 'errors': serializer.errors})
        if errors:
            return Response({'correct_contacts': contacts, 'errors': errors},
                            status=status.HTTP_400_BAD_REQUEST)
        serializer = CreateContactSerializer(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({'contacts': contacts},
                        status=status.HTTP_201_CREATED)

    def check_email_in_request(self, id: int, contact_data: Dict,
                               contacts: List[Dict]) -> bool:
        checked_email: str = contact_data['email']
        for previous_contact in contacts[:id]:
            if previous_contact['email'] == checked_email:
                return False
        return True

class DeleteContactView(DestroyAPIView):
    serializer_class = SimpleContactSerializer
    queryset = Contact.active_objects.all()
    lookup_field = 'id'
    
    @swagger_auto_schema(operation_id='delete_contact')
    def delete(self, reuqest: Request, id: str) -> Response:
        contact: Contact = self.queryset.filter(id=id).first()
        if not contact:
            return Response({'error': ErrorMessages.NO_CONTACT.value},
                            status=status.HTTP_404_NOT_FOUND)
        contact.soft_delete()
        return Response({'message': 'Deletion complete!'},
                        status=status.HTTP_204_NO_CONTENT)


class UpdateContactView(GenericAPIView):
    serializer_class = UpdateContactSerializer
    queryset = Contact.active_objects.all()
    # permission_classes  = (IsAdminUser,)

    @swagger_auto_schema(operation_id='update_contact')
    def put(self, request: Request, id: str) -> Response:
        contact: Contact = self.queryset.filter(id=id).first()
        if not contact:
            return Response({'error': ErrorMessages.NO_CONTACT.value},
                            status=status.HTTP_404_NOT_FOUND)

        serialilizer = self.serializer_class(
            instance=contact, data=request.data, partial=True)
        serialilizer.is_valid(raise_exception=True)
        serialilizer.save()
        return Response(serialilizer.data, status=status.HTTP_200_OK)


class UploadContactsSheetView(APIView):
    parser_classes = (MultiPartParser,)
    serializer_class = None

    @swagger_auto_schema(
        operation_id='upload_sheet',
        manual_parameters=[
            openapi.Parameter(
                'file',
                in_=openapi.IN_FORM,
                type=openapi.TYPE_FILE,
                required=True,
                description='Excel sheet file'
            ),
        ]
    )
    def post(self, request: Request) -> Response:
        file_object: File = request.data['file']
        file_name = file_object.name
        file_extension = os.path.splitext(file_name)[1].lower()

        if file_extension != '.xlsx':
            return Response({'error': ErrorMessages.INVALID_TYPE.value},
                            status=status.HTTP_400_BAD_REQUEST)

        wb = load_workbook(file_object)
        sheet: Worksheet = wb.active
        contacts = []
        errors = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, email, phone, priority_group = row

            contact_data = {
                'name': name,
                'email': email,
                'phone': phone,
                'priority_group': priority_group,
            }

            serializer = CreateContactSerializer(data=contact_data)
            if serializer.is_valid():
                serializer.save()
                contacts.append(serializer.data)
            else:
                errors.append(serializer.errors)

        return Response({'contacts': contacts, 'errors': errors}, status=status.HTTP_200_OK)